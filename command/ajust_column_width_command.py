import openpyxl
from RendingDTO import RendingDataSet
from IExecutable import IExecutable
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

#
#
#


class AdjustColumnWidthCommand(IExecutable):

    SHEET_NAME = "Rending Ratio"
    TARGET_COLUMN = "D"
    TARGET_WIDTH = 17.73

    def execute(self, dto: RendingDataSet) -> RendingDataSet:

        sheet = dto.excel_workbook[self.SHEET_NAME]
        # D列の幅を設定、以下、個別の列サイズなのでマジックナンバーOKとする
        sheet.column_dimensions["A"].width = 28.09
        sheet.column_dimensions["D"].width = 17.73
        sheet.column_dimensions["E"].width = 15.45
        sheet.column_dimensions["F"].width = 14.00
        sheet.column_dimensions["G"].width = 14.00
        sheet.column_dimensions["H"].width = 24.00
        sheet.column_dimensions["I"].width = 19.09
        sheet.column_dimensions["J"].width = 18.00
        sheet.column_dimensions["K"].width = 14.00
        sheet.column_dimensions["L"].width = 19.27
        sheet.row_dimensions[1].height = 26
        sheet.cell(row=1, column=8).alignment = Alignment(wrap_text=True)
        sheet.cell(row=1, column=9).alignment = Alignment(wrap_text=True)
        sheet.cell(row=1, column=11).alignment = Alignment(wrap_text=True)

        # sheet.auto_filter.ref = sheet.dimensions
        # sheet.auto_filter.add_sort_condition("G2:G{}".format(sheet.max_row), descending=True)

        # sheet.sort('G')

        # ワークブックを保存
        dto.excel_workbook.save(dto.out_filepath)
        return dto
