import os
import openpyxl
from RendingDTO import RendingDataSet
from IExecutable import IExecutable
from datetime import datetime


class WriteWeeklyDataToExcelCommand(IExecutable):
    def execute(self, dto: RendingDataSet) -> RendingDataSet:
        extracted_lines = dto.weekly_outstanding_data

        # ファイル名に年月日を追加
        current_date = datetime.now().strftime("%Y%m%d")
        output_file_path = f'weekly_data_{current_date}.xlsx'

        # 既に同名のファイルが存在する場合、別名で保存する
        counter = 1
        while os.path.exists(output_file_path):
            output_file_name = f'weekly_data_{current_date}_{counter}.xlsx'
            output_file_path = output_file_name
            counter += 1

        # Excelファイルにデータを書き込む
        dto.out_filepath = output_file_path
        dto.excel_workbook = openpyxl.Workbook()
        worksheet = dto.excel_workbook.create_sheet(title="JPX Data")
        # worksheet = workbook.active

        for line in extracted_lines:
            row_data = line.split(",")
            converted_data = []
            for value in row_data:
                try:
                    converted_value = int(value)
                except ValueError:
                    converted_value = value
                converted_data.append(converted_value)
            worksheet.append(converted_data)

        worksheet.freeze_panes = 'A2'
        dto.jpx_worksheet = worksheet

        # ファイルを保存する前にデフォルトの "Sheet" シートを削除する
        if 'Sheet' in dto.excel_workbook.sheetnames:
            dto.excel_workbook.remove(dto.excel_workbook['Sheet'])

        # workbook.save(output_file_path)

        return dto
