import os
import openpyxl
from RendingDTO import RendingDataSet
from IExecutable import IExecutable
from datetime import datetime


class WriteWeeklyRatioDataToExcelCommand:
    def execute(self, dto: RendingDataSet) -> RendingDataSet:
        extracted_lines = dto.weekly_outstanding_data

        rending_ratio_worksheet = dto.excel_workbook.create_sheet(title="Rending Ratio")

        headers = [
            "銘柄名", "コード", "売り残高＋貸株残高", "浮動株数", "発行済み株数", "売り残高＋貸株残高/浮動株数", "売り残高＋貸株残/発行済み株数", "売り残高＋貸株残高前週比", "売り残高＋貸株残高/買残高", "売残高", "売残高前週比", "貸株残高", "貸株残高前週比",
            "買残高", "買残高前週比", "一般信用売残高", "一般信用売残高前週比", "制度信用売残高", "制度信用売残高前週比",
            "一般信用買残高", "一般信用買残高前週比", "制度信用買残高", "一般信用買残高前週比"
        ]
        rending_ratio_worksheet.append(headers)

        code_column = "B"  # コードが格納されている列のアルファベットを指定
        #code_column = dto.jpx_worksheet.cell(row=1, column=1).column_letter
        #print("dto.stock_list[1301])

        for row_index, row in enumerate(dto.jpx_worksheet.iter_rows(min_row=2, values_only=True), start=2):
            code_cell = dto.jpx_worksheet[code_column + str(row_index)]
            code = code_cell.value
            shortcode = str(code_cell.value)[0:4]
            total_sales_and_lending = row[3]
            total_outstanding_purchases = row[3]
            diff_sales_and_lending = row[4]
            total_lending = 0
            diff_lending = 0

            for nisshokyo_row in dto.nisshokyo_worksheet.iter_rows(min_row=2):
                if nisshokyo_row[1].value == code:
                    # total_outstanding_purchases += nisshokyo_row[3].value
                    # total_sales_and_lending += nisshokyo_row[4].value + nisshokyo_row[5].value
                    total_sales_and_lending += nisshokyo_row[3].value
                    total_lending += nisshokyo_row[3].value
                    #total_outstanding_purchases = nisshokyo_row[3].value
                    # diff_sales_and_lending = row[4] + int(nisshokyo_row[4].value.replace('▲,', '-'))
                    # diff_lending = int(nisshokyo_row[4].value.replace('▲,', '-'))
                    if nisshokyo_row[4].value != '-':
                        diff_lending += int(nisshokyo_row[4].value)

                    if nisshokyo_row[4].value != '-':
                        diff_lending += int(nisshokyo_row[4].value)

            # print(row_index)
            # print(code)
            if row[5] == 0:
                ratio_sales_and_lending_purchases = 0
            else:
                ratio_sales_and_lending_purchases = total_sales_and_lending/row[5]

            # print(row[1])
            # print(row[2])
            # print(row[3])
            # print(row[4])
            # print(total_sales_and_lending)
            # print(total_lending)
            # print(diff_sales_and_lending)
            # print(diff_lending)
            # print(dto.stock_list[int(1301)])

            if float(dto.stock_list[int(shortcode)].stock_float) == 0:
                total_sales_and_lending_float_ratio = 0
            else:
                total_sales_and_lending_float_ratio = total_sales_and_lending / float(dto.stock_list[int(shortcode)].stock_float)

            if float(dto.stock_list[int(shortcode)].stock_shares_outstanding) == 0:
                total_sales_and_lending_outstanding_ratio = 0
            else:
                total_sales_and_lending_outstanding_ratio = row[5] / float(dto.stock_list[int(shortcode)].stock_shares_outstanding)

            row_data = [
                dto.stock_list[int(shortcode)].name,  # 銘柄名
                dto.stock_list[int(shortcode)].code,
                total_sales_and_lending,  # 売り残高＋貸株残高
                dto.stock_list[int(shortcode)].stock_float, # 浮動株数
                dto.stock_list[int(shortcode)].stock_shares_outstanding, #発行済み株数
                total_sales_and_lending_float_ratio,  # "売り残高＋貸株残高/浮動株数"
                total_sales_and_lending_outstanding_ratio,  # "売り残高＋貸株残高/浮動株数"
                diff_sales_and_lending,  # 売り残高＋貸株残高先週比
                ratio_sales_and_lending_purchases, # "売り残高＋貸株残高/買残高"
                row[3],  # 売残高
                row[4],  # 売残高前週比
                total_lending,  # 貸株残高
                diff_lending,  # 貸株残高前週比
                row[5],  # 買残高
                row[6],  # 買残高前週比
                row[7],  # 一般信用売残高
                row[8],  # 一般信用売残高前週比
                row[9],  # 制度信用売残高
                row[10],  # 制度信用売残高前週比
                row[11],  # 一般信用買残高
                row[12],  # 一般信用買残高前週比
                row[13],  # 制度信用買残高
                row[14]   # 一般信用買残高前週比
            ]

            rending_ratio_worksheet.append(row_data)
        rending_ratio_worksheet.freeze_panes = 'A2'

        # ファイルを保存する前にデフォルトの "Sheet" シートを削除する
        if 'Sheet' in dto.excel_workbook.sheetnames:
            dto.excel_workbook.remove(dto.excel_workbook['Sheet'])

        dto.excel_workbook.save(dto.out_filepath)

        dto.rending_ratio_worksheet = rending_ratio_worksheet

        return dto
